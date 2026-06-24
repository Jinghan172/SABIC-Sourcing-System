"""
一键导出所有品类供应商总表
用法：python export_all.py
输出：SABIC_供应商总表_YYYYMMDD.xlsx
"""
import sys, json, datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent / "sabic-py"))
sys.path.insert(0, str(Path(__file__).parent))
from utils.local_search import search_local, CACHE_DIR
from collect_local import CATEGORIES

# ── 品类顺序与分组：直接取自 collect_local.py 的 147 个品类 ────────────
GROUPS = {grp: [(en, cn) for en, cn, _kw in items]
          for grp, items in CATEGORIES.items()}

ROLE_ZH = {
    "manufacturer": "工厂",
    "both":         "工厂兼贸易",
    "importer":     "进口商",
    "trader":       "经销商",
    "agent":        "中介",
    "unknown":      "未分类",
}

TIER_ZH = {1: "一级(华东)", 2: "二级", 3: "三级"}

# ── 颜色主题 ─────────────────────────────────────────────────────────
C_HEADER    = "0A1628"   # 深蓝标题行
C_GRP_BG    = {          # 分组底色（淡色）
    "原材料":  "EFF6FF",
    "阻燃剂":  "FFF7ED",
    "改性剂":  "F0FDF4",
    "氟塑料":  "ECFEFF",
    "稳定剂":  "FDF4FF",
    "色料增强":"FEFCE8",
    "包装":    "F1F5F9",
}
C_GRP_FONT  = {
    "原材料":  "1D4ED8",
    "阻燃剂":  "C2410C",
    "改性剂":  "15803D",
    "氟塑料":  "0E7490",
    "稳定剂":  "7E22CE",
    "色料增强":"A16207",
    "包装":    "475569",
}
C_SCORE_HIGH = "D1FAE5"
C_SCORE_MID  = "FEF3C7"
C_SCORE_LOW  = "FEE2E2"
C_WHITE      = "FFFFFF"
C_ALTROW     = "F8FAFC"

def _fill(hex6): return PatternFill("solid", fgColor=hex6)
def _font(hex6="000000", bold=False, sz=10):
    return Font(name="微软雅黑", color=hex6, bold=bold, size=sz)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

COLUMNS = [
    ("分组",        10),
    ("品类(EN)",     9),
    ("品类(中文)",   10),
    ("综合评分",      8),
    ("排名",          6),
    ("企业名称",     28),
    ("企业类型",      9),
    ("地理评分",      8),
    ("规模评分",      8),
    ("合规评分",      8),
    ("圈层",          8),
    ("省份",          7),
    ("城市",          7),
    ("注册资本",     11),
    ("成立年份",      8),
    ("经营状态",      8),
    ("危化品资质",    8),
    ("统一信用代码",  20),
    ("法定代表人",    10),
    ("地址",         35),
    ("经营范围(摘要)", 50),
]

def _score_color(score):
    if score >= 65: return C_SCORE_HIGH
    if score >= 45: return C_SCORE_MID
    return C_SCORE_LOW

def collect_all_rows():
    rows = []
    for grp, items in GROUPS.items():
        for en, cn in items:
            result = search_local(cn, top_n=50)
            suppliers = result.get("suppliers", [])
            for rank, s in enumerate(suppliers, 1):
                d = s.get("dimensions", {})
                lic = s.get("licenses", {})
                scope = s.get("_business_scope", "") or ""
                rows.append({
                    "grp":     grp,
                    "en":      en,
                    "cn":      cn,
                    "score":   s.get("score", 0),
                    "rank":    rank,
                    "name":    s.get("name", ""),
                    "role":    ROLE_ZH.get(s.get("_role","unknown"), "未分类"),
                    "geo":     d.get("geography", 0),
                    "scale":   d.get("scale", 0),
                    "comp":    d.get("compliance", 0),
                    "tier":    TIER_ZH.get(s.get("_tier", 3), "三级"),
                    "province":s.get("province", ""),
                    "city":    s.get("city", ""),
                    "capital": (f"{s.get('registered_capital_wan',0)/10000:.1f}亿"
                                if (s.get("registered_capital_wan") or 0) >= 10000
                                else f"{s.get('registered_capital_wan',0):.0f}万"
                                if s.get("registered_capital_wan") else ""),
                    "est":     s.get("established", "") or "",
                    "status":  s.get("reg_status", "存续") or "存续",
                    "hazmat":  "✓" if (lic.get("hazardous_chemicals") or
                                       lic.get("hazmat_business")) else "",
                    "credit":  s.get("creditCode","") or s.get("credit_code","") or "",
                    "legal":   s.get("legalPerson","") or s.get("legal_person","") or "",
                    "addr":    s.get("address","") or "",
                    "scope":   scope[:200] + ("…" if len(scope) > 200 else ""),
                })
    return rows

def build_excel(rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "供应商总表"
    ws.freeze_panes = "A3"

    # ── 第一行：大标题 ─────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value = (f"SABIC 上海 · 化工原材料供应商总表 "
                        f"（共 {len(rows)} 家 · 生成于 "
                        f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}）")
    title_cell.fill    = _fill(C_HEADER)
    title_cell.font    = _font("FFFFFF", bold=True, sz=12)
    title_cell.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # ── 第二行：列头 ──────────────────────────────────────────────
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=col_idx, value=name)
        c.fill      = _fill("1E3A5F")
        c.font      = _font("FFFFFF", bold=True)
        c.alignment = _align("center")
        c.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 20

    # ── 数据行 ────────────────────────────────────────────────────
    prev_grp = None
    for row_idx, r in enumerate(rows, 3):
        grp = r["grp"]
        bg = C_GRP_BG.get(grp, C_WHITE) if r["rank"] == 1 else (
            C_ALTROW if row_idx % 2 == 0 else C_WHITE)
        score_bg = _score_color(r["score"])

        vals = [
            r["grp"], r["en"], r["cn"],
            round(r["score"], 1), r["rank"],
            r["name"], r["role"],
            round(r["geo"], 1), round(r["scale"], 1),
            round(r["comp"], 1),
            r["tier"], r["province"], r["city"],
            r["capital"], r["est"], r["status"], r["hazmat"],
            r["credit"], r["legal"], r["addr"],
            r["scope"],
        ]

        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border    = _thin_border()
            c.alignment = _align("left" if col_idx > 5 else "center",
                                  wrap=(col_idx == len(COLUMNS)))
            c.font = _font()

            # 评分列用颜色编码
            if col_idx == 4:  # 综合评分
                c.fill = _fill(score_bg)
                c.font = _font("0E8C3A" if r["score"] >= 65 else
                               "D97706" if r["score"] >= 45 else "DC2626",
                               bold=True)
                c.alignment = _align("center")
            elif col_idx in (8, 9, 10):  # 三维评分
                c.fill = _fill(score_bg)
                c.alignment = _align("center")
            elif col_idx == 1 and r["rank"] == 1:  # 分组第一家加色
                c.fill = _fill(C_GRP_BG.get(grp, C_WHITE))
                c.font = _font(C_GRP_FONT.get(grp, "000000"), bold=True)
            else:
                c.fill = _fill(bg)

            # 状态列：注销/吊销标红
            if col_idx == 16 and val not in ("存续", "在业", ""):
                c.font = _font("DC2626", bold=True)

        ws.row_dimensions[row_idx].height = 16

    # ── 按分组+分数排序已在 collect_all_rows 里完成 ───────────────

    # ── 第二个 Sheet：品类汇总 ─────────────────────────────────────
    ws2 = wb.create_sheet("品类汇总")
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 10

    hdr2 = ["分组", "品类(EN)", "品类(中文)", "供应商数", "平均总分", "最高总分"]
    for ci, h in enumerate(hdr2, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill = _fill("1E3A5F"); c.font = _font("FFFFFF", bold=True)
        c.alignment = _align("center"); c.border = _thin_border()
    ws2.row_dimensions[1].height = 20

    from collections import defaultdict
    summary = defaultdict(list)
    for r in rows:
        summary[(r["grp"], r["en"], r["cn"])].append(r["score"])

    prev_grp2 = None
    ri2 = 2
    for (grp, en, cn), scores in summary.items():
        n = len(scores)
        avg = round(sum(scores)/n, 1) if n else 0
        mx  = round(max(scores), 1) if n else 0
        bg2 = C_GRP_BG.get(grp, C_WHITE)
        vals2 = [grp, en, cn, n, avg, mx]
        for ci, v in enumerate(vals2, 1):
            c = ws2.cell(row=ri2, column=ci, value=v)
            c.fill = _fill(bg2); c.font = _font()
            c.alignment = _align("center"); c.border = _thin_border()
            if ci == 4 and v == 0:
                c.font = _font("9CA3AF")
        ws2.row_dimensions[ri2].height = 16
        ri2 += 1

    wb.save(out_path)
    return len(rows)

if __name__ == "__main__":
    print("正在读取所有品类数据并评分...")
    rows = collect_all_rows()
    date_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    out = Path(__file__).parent / f"SABIC_供应商总表_{date_str}.xlsx"
    n = build_excel(rows, out)
    print(f"[OK] 导出完成：{out}")
    print(f"  共 {n} 家供应商（含全部品类）")
