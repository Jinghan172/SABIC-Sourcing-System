# -*- coding: utf-8 -*-
"""
SABIC 供应商总表导出器 —— 一次导出全站所有品类的供应商，多 Sheet 工作簿。

覆盖网页上的四类来源：
  ①  专家评审 · 核心物料   (core_materials.json，6 维专家评分，含国产+进口)
  ②  核心采购品类 P1        (local_cache/*.json，企查查 3 维工商评分)
  ③  补充扩展品类 P2        (local_cache/*.json，企查查 3 维工商评分)
  ④  综合服务 · 15 品类     (services.json，5 维加权专家评分，4 基地 ×5 家)

用法：  python export_master.py
输出：  ../SABIC_供应商总表_全品类_<时间戳>.xlsx
"""
from __future__ import annotations
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.local_search import list_cache_categories, search_local
from utils.sabic_search import get_category_priority
from utils.services_scorer import rank_suppliers, DIM_KEYS as SV_DIMS, DIM_CN as SV_DIM_CN, \
    DEFAULT_WEIGHTS as SV_DEFAULT_W, verdict_for

APP_DIR = Path(__file__).resolve().parent
DATA = APP_DIR / "data"

SABIC_GREEN = "FF0E8C3A"
SABIC_DARK = "FF0A1628"
SABIC_BLUE = "FF2563EB"
WHITE = "FFFFFFFF"
LIGHT_GREEN = "FFF0FAF4"
LIGHT_GRAY = "FFF6F8FB"
_THIN = Side(style="thin", color="FFD9E1EC")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _title(ws, ncols, text):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=f"SABIC 上海寻源系统 · {text}")
    cell.font = Font(name="微软雅黑", bold=True, size=13, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=SABIC_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


def _headers(ws, headers, color=SABIC_DARK, row=2):
    for col, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(name="微软雅黑", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 26


def _score_font(v):
    try:
        v = float(v)
    except (TypeError, ValueError):
        return Font(name="微软雅黑", size=9)
    if v >= 70:
        return Font(name="微软雅黑", size=9, bold=True, color="FF059669")
    if v >= 50:
        return Font(name="微软雅黑", size=9, color="FFD97706")
    return Font(name="微软雅黑", size=9, color="FFDC2626")


def _write_rows(ws, rows, start_row=3, score_cols=(), name_cols=()):
    for i, vals in enumerate(rows):
        r = start_row + i
        even = i % 2 == 1
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name="微软雅黑", size=9)
            cell.border = BORDER
            if col in name_cols:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if even:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            if col in score_cols:
                cell.font = _score_font(val)
        ws.row_dimensions[r].height = 17
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(rows[0]) if rows else 1)}2"


TIER_LABEL = ["", "一级(华东)", "二级", "三级"]
ROLE_ZH = {"manufacturer": "工厂", "both": "工厂兼贸易", "importer": "进口商",
           "trader": "经销商", "agent": "中介", "unknown": "未分类"}
BASE_CN = {"SH": "上海浦东", "NS": "广州南沙", "GL": "福建古雷", "CQ": "重庆"}
NATURE_ZH = {"foreign": "外资", "soe": "国企", "joint": "合资", "private": "民营"}
TIERN_ZH = {"national_top": "全国龙头", "regional": "区域龙头", "local": "属地本地"}
SECTOR_ZH = {"petrochem": "石化炼化", "chemical": "化工", "industrial": "工业通用", "general": "通用"}


# ════════════════════════════════════════════════════════════════════
# Sheet ① 专家评审 · 核心物料（6 维）
# ════════════════════════════════════════════════════════════════════
def sheet_expert(wb, stats):
    cm = json.loads((DATA / "core_materials.json").read_text(encoding="utf-8"))
    dim_keys = cm.get("dim_keys", ["tech", "location", "cost", "ehss", "scale", "service"])
    dim_cn = cm.get("dim_cn", {})
    ws = wb.create_sheet("① 专家评审·核心物料")
    ws.sheet_properties.tabColor = "0E8C3A"
    heads = [("核心物料", 13), ("国产/进口", 9), ("排名", 6), ("供应商全称", 30),
             ("综合评分", 9)] + [(dim_cn.get(k, k), 11) for k in dim_keys] + \
            [("所在地/产地", 22), ("类型", 16), ("工商/联网核验", 30),
             ("专家评语", 18)]
    _title(ws, len(heads), "供应商总表 · ① 专家评审核心物料（6 维差异化加权专家评分）")
    _headers(ws, heads, color=SABIC_GREEN)
    rows = []
    n = 0
    for m in cm.get("materials", []):
        seq = [(c, "国产") for c in m.get("companies", [])] + \
              [(c, "进口") for c in m.get("import_companies", [])]
        for c, origin in seq:
            d = c.get("dims", {})
            rows.append([m.get("cn", ""), origin, c.get("rank", ""), c.get("name", ""),
                         c.get("score", "")] + [round(d.get(k, 0), 1) for k in dim_keys] +
                        [c.get("location", ""), c.get("type", ""), c.get("gs", ""),
                         c.get("verdict", "")])
            n += 1
    score_cols = tuple([5] + list(range(6, 6 + len(dim_keys))))
    _write_rows(ws, rows, score_cols=score_cols, name_cols=(4, 7, 9, 10))
    stats.append(("① 专家评审 · 核心物料", f"{len(cm.get('materials', []))} 类战略物料", n))


# ════════════════════════════════════════════════════════════════════
# Sheet ②③ 工商评分品类（3 维）—— 核心 P1 / 补充 P2
# ════════════════════════════════════════════════════════════════════
def _gongshang_sheet(wb, title, want_priority, tab_color, stats, label):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color
    heads = [("采购品类", 16), ("排名", 6), ("供应商全称", 30), ("综合评分", 9),
             ("地理评分", 9), ("规模评分", 9), ("合规资质", 9), ("省份", 8),
             ("城市", 9), ("地理圈层", 11), ("企业类型", 11), ("经营状态", 9),
             ("注册资本(万)", 12), ("成立年", 8), ("危化品资质", 10), ("化工园区", 9)]
    _title(ws, len(heads), f"供应商总表 · {label}（企查查工商数据 · 地理/规模/合规 三维评分）")
    _headers(ws, heads, color=(SABIC_DARK if want_priority == 1 else SABIC_BLUE))
    rows = []
    ncat = 0
    for cat in list_cache_categories():
        pr = get_category_priority(cat["cn"])
        is_p2 = (pr == 2)
        if want_priority == 2 and not is_p2:
            continue
        if want_priority == 1 and is_p2:
            continue
        res = search_local(cat["cn"], filters={}, top_n=9999)
        sups = res.get("suppliers", [])
        if not sups:
            continue
        ncat += 1
        for i, s in enumerate(sups):
            dims = s.get("dimensions", {})
            lic = s.get("licenses", {})
            rows.append([
                cat["cn"], i + 1, s.get("name", ""), s.get("score", 0),
                round(dims.get("geography", 0), 1), round(dims.get("scale", 0), 1),
                round(dims.get("compliance", 0), 1), s.get("province", ""),
                s.get("city", ""), TIER_LABEL[s.get("_tier", 3)],
                ROLE_ZH.get(s.get("_role", "unknown"), "未分类"),
                s.get("reg_status", "存续") or "存续",
                s.get("registered_capital_wan", ""), s.get("established", "") or "",
                "✓" if (lic.get("hazardous_chemicals") or lic.get("hazmat_business")) else "—",
                "✓" if s.get("chemical_park") else "—",
            ])
    _write_rows(ws, rows, score_cols=(4, 5, 6, 7), name_cols=(3,))
    stats.append((label, f"{ncat} 个品类", len(rows)))


# ════════════════════════════════════════════════════════════════════
# Sheet ④ 综合服务 · 15 品类（5 维）
# ════════════════════════════════════════════════════════════════════
def sheet_services(wb, stats):
    sv = json.loads((DATA / "services.json").read_text(encoding="utf-8"))
    ws = wb.create_sheet("④ 综合服务·15品类")
    ws.sheet_properties.tabColor = "7C3AED"
    heads = [("服务品类", 16), ("基地", 11), ("排名", 6), ("供应商全称", 30),
             ("综合评分", 9)] + [(SV_DIM_CN[k], 11) for k in SV_DIMS] + \
            [("角色", 9), ("规模圈层", 10), ("企业性质", 9), ("行业适配", 10),
             ("类型", 14), ("资质/合规标签", 24), ("专家评语", 18)]
    _title(ws, len(heads), "供应商总表 · ④ 综合服务 15 品类（5 维加权专家评分 · 四大基地 ×5 家）")
    _headers(ws, heads, color="FF7C3AED")
    rows = []
    for c in sv.get("categories", []):
        w = c.get("weights", SV_DEFAULT_W)
        for bk, bcn in BASE_CN.items():
            sl = c.get("bases", {}).get(bk, {}).get("suppliers", [])
            for s in rank_suppliers(sl, w):
                d = s["dims"]
                rows.append([c["cn"], bcn, s["rank"], s["name"], s["score"]] +
                            [round(d[k], 1) for k in SV_DIMS] +
                            ["采购推荐首选" if s.get("role") == "primary" else "备选",
                             TIERN_ZH.get(s.get("tier", ""), ""),
                             NATURE_ZH.get(s.get("nature", ""), ""),
                             SECTOR_ZH.get(s.get("sector", ""), ""),
                             s.get("type", ""), "、".join(s.get("quals", [])),
                             verdict_for(s)])
    score_cols = tuple([5] + list(range(6, 6 + len(SV_DIMS))))
    _write_rows(ws, rows, score_cols=score_cols, name_cols=(4, 13, 14, 15))
    stats.append(("④ 综合服务 · 15 品类", "15 品类 × 4 基地", len(rows)))


# ════════════════════════════════════════════════════════════════════
# Sheet 0 汇总封面
# ════════════════════════════════════════════════════════════════════
def sheet_cover(wb, stats):
    ws = wb.create_sheet("汇总", 0)
    ws.sheet_properties.tabColor = "0A1628"
    _title(ws, 4, "供应商总表 · 全品类汇总")
    sub = ws.cell(row=2, column=1,
                  value=f"导出时间：{datetime.now():%Y-%m-%d %H:%M}　|　"
                        f"覆盖：专家评审核心物料 + 核心采购品类 + 补充扩展品类 + 综合服务 15 品类")
    ws.merge_cells("A2:D2")
    sub.font = Font(name="微软雅黑", size=10, color="FF334155")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 24
    _headers(ws, [("分区 / 数据来源", 34), ("评分模型", 18), ("覆盖范围", 20),
                  ("供应商数量", 14)], color=SABIC_DARK, row=4)
    models = {
        "① 专家评审 · 核心物料": "6 维差异化专家评分",
        "② 核心采购品类 P1": "企查查 3 维工商评分",
        "③ 补充扩展品类 P2": "企查查 3 维工商评分",
        "④ 综合服务 · 15 品类": "5 维加权专家评分",
    }
    total = 0
    for i, (name, scope, n) in enumerate(stats):
        r = 5 + i
        total += n
        for col, val in enumerate([name, models.get(name, ""), scope, n], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name="微软雅黑", size=10,
                             bold=(col == 4))
            cell.alignment = Alignment(horizontal=("left" if col == 1 else "center"),
                                       vertical="center")
            cell.border = BORDER
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        ws.row_dimensions[r].height = 22
    tr = 5 + len(stats)
    for col, val in enumerate(["合计", "", "全站全品类", total], 1):
        cell = ws.cell(row=tr, column=col, value=val)
        cell.font = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=SABIC_GREEN)
        cell.alignment = Alignment(horizontal=("left" if col == 1 else "center"),
                                   vertical="center")
        cell.border = BORDER
    ws.row_dimensions[tr].height = 26


def build_master() -> bytes:
    import io
    wb = Workbook()
    wb.remove(wb.active)  # 删掉默认空 sheet
    stats = []
    sheet_expert(wb, stats)
    _gongshang_sheet(wb, "② 核心采购品类", 1, "0A1628", stats, "② 核心采购品类 P1")
    _gongshang_sheet(wb, "③ 补充扩展品类", 2, "2563EB", stats, "③ 补充扩展品类 P2")
    sheet_services(wb, stats)
    sheet_cover(wb, stats)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    data = build_master()
    out = APP_DIR.parent / f"SABIC_供应商总表_全品类_{datetime.now():%Y%m%d_%H%M}.xlsx"
    out.write_bytes(data)
    print(f"OK · 已导出：{out}  （{len(data) / 1024:.0f} KB）")
