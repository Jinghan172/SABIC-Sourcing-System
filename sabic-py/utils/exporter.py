"""Excel 导出 — 使用 openpyxl，返回 bytes 供 Streamlit 下载按钮使用"""
from __future__ import annotations
import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

SABIC_GREEN = "FF0E8C3A"
SABIC_DARK  = "FF0A1628"
WHITE       = "FFFFFFFF"
LIGHT_GREEN = "FFF0FAF4"
LIGHT_GRAY  = "FFF2F5F9"


def _header_style(cell, dark: bool = True):
    cell.font = Font(name="微软雅黑", bold=True, color=WHITE, size=10)
    bg = SABIC_DARK if dark else SABIC_GREEN
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _data_style(cell, even: bool = False):
    cell.font = Font(name="微软雅黑", size=9)
    if even:
        cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def export_excel(suppliers: list[dict], title: str = "供应商对比") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "供应商评分"
    ws.sheet_properties.tabColor = "0E8C3A"

    # ── 标题行 ────────────────────────────────────────────────────────
    ws.merge_cells("A1:R1")
    ws["A1"] = f"SABIC 上海寻源系统 · {title}"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=13, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=SABIC_DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── 列头 ─────────────────────────────────────────────────────────
    headers = [
        ("排名", 6), ("供应商简称", 14), ("供应商全称", 24),
        ("综合评分", 9), ("地理评分", 9),
        ("规模评分", 9), ("合规资质", 9),
        ("所在省份", 9), ("城市", 9), ("地理圈层", 10),
        ("企业类型", 10), ("经营状态", 9),
        ("注册资本(万)", 12), ("成立年份", 10),
        ("危化品资质", 10), ("安全生产证", 10), ("化工园区", 9),
        ("备注", 16),
    ]
    for col, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    # ── 数据行 ────────────────────────────────────────────────────────
    def yn(val: bool) -> str:
        return "✓" if val else "—"

    ROLE_ZH = {"manufacturer": "工厂", "both": "工厂兼贸易", "importer": "进口商",
               "trader": "经销商", "agent": "中介", "unknown": "未分类"}

    for i, s in enumerate(suppliers):
        row = i + 3
        dims = s.get("dimensions", {})
        lic  = s.get("licenses", {})
        tier = s.get("_tier", 3)
        tier_label = ["", "一级(华东)", "二级", "三级"][tier]

        values = [
            i + 1,
            s.get("shortName", s.get("name", "")[:8]),
            s.get("name", ""),
            s.get("score", 0),
            round(dims.get("geography", 0), 1),
            round(dims.get("scale", 0), 1),
            round(dims.get("compliance", 0), 1),
            s.get("province", ""),
            s.get("city", ""),
            tier_label,
            ROLE_ZH.get(s.get("_role", "unknown"), "未分类"),
            s.get("reg_status", "存续") or "存续",
            s.get("registered_capital_wan", ""),
            s.get("established", ""),
            yn(lic.get("hazardous_chemicals") or lic.get("hazmat_business")),
            yn(lic.get("safety_production")),
            yn(s.get("chemical_park")),
            "",
        ]

        even = (i % 2 == 1)
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _data_style(cell, even)
            # 评分用条件格式色 (简单实现：高分加绿色)
            if col == 4 and isinstance(val, (int, float)):
                if val >= 70:
                    cell.font = Font(name="微软雅黑", size=9,
                                     color="FF059669", bold=True)
                elif val >= 50:
                    cell.font = Font(name="微软雅黑", size=9, color="FFD97706")
                else:
                    cell.font = Font(name="微软雅黑", size=9, color="FFDC2626")

        ws.row_dimensions[row].height = 18

    # ── 冻结首两行 + 筛选 ─────────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
